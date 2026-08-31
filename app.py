import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from collections import Counter
import io

# ── Configuração da página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Previne Niterói — SMF",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Estilo ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
  html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

  /* cabeçalho */
  .header-bar {
    background: linear-gradient(135deg, #0d3b6e 0%, #1565c0 100%);
    color: white;
    padding: 20px 28px;
    border-radius: 12px;
    margin-bottom: 24px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    flex-wrap: wrap;
    gap: 12px;
  }
  .header-title { font-size: 1.15rem; font-weight: 700; margin-bottom: 4px; }
  .header-sub   { font-size: .78rem; opacity: .75; }
  .header-badge {
    background: rgba(255,255,255,.15);
    border-radius: 20px;
    padding: 4px 14px;
    font-size: .78rem;
    white-space: nowrap;
  }
  .header-meta  { font-size: .72rem; opacity: .75; text-align: right; }

  /* KPI cards */
  .kpi-row { display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 20px; }
  .kpi {
    background: white;
    border-radius: 10px;
    padding: 14px 18px;
    flex: 1;
    min-width: 150px;
    box-shadow: 0 1px 4px rgba(0,0,0,.07);
    border-top: 4px solid #94a3b8;
  }
  .kpi .val { font-size: 2rem; font-weight: 800; line-height: 1.1; }
  .kpi .lbl { font-size: .72rem; color: #94a3b8; margin-top: 4px; }

  /* alerta */
  .alerta {
    background: #fcebe0;
    border: 1px solid #f3c9ae;
    border-radius: 10px;
    padding: 12px 16px;
    font-size: .88rem;
    margin-bottom: 20px;
  }
  .alerta b { color: #e8611a; }

  /* owner bar */
  .owner {
    background: white;
    border-left: 4px solid #1565c0;
    border-radius: 10px;
    padding: 12px 18px;
    margin-bottom: 20px;
    box-shadow: 0 1px 4px rgba(0,0,0,.07);
  }
  .owner .lbl { font-size: .68rem; font-weight: 700; text-transform: uppercase;
                 letter-spacing: .5px; color: #94a3b8; margin-bottom: 3px; }
  .owner .val { font-size: 1rem; font-weight: 700; color: #0d3b6e; }
  .owner .desc{ font-size: .78rem; color: #64748b; margin-top: 3px; }

  /* tabelas */
  .stDataFrame { border-radius: 10px; overflow: hidden; }
  div[data-testid="stDataFrameContainer"] { border-radius: 10px; }

  /* esconder menu e rodapé padrão */
  #MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ── Constantes ───────────────────────────────────────────────────────────────
STATUS_COLORS = {
    "Concluído":    "#2e7d32",
    "Em andamento": "#e8a100",
    "Contínuo":     "#1565c0",
    "Não iniciado": "#c62828",
}
SERIE_HIST = {
    "jan/26": 47, "fev/26": 47, "mar/26": 50,
    "abr/26": 50, "mai/26": 50, "jun/26": 52,
    "jul/26": 58,
}

# ── Leitura da planilha ──────────────────────────────────────────────────────
def ler_planilha(fileobj):
    wb = load_workbook(fileobj, data_only=True)
    ws = wb["Previne 2025-2026"]
    rows = []
    for r in range(15, 120):
        a = ws.cell(row=r, column=1).value
        try:
            n = int(str(a).strip())
        except:
            continue
        p    = str(ws.cell(row=r, column=3).value or "").replace("PREVINE - PILAR ", "").strip()
        resp = " ".join(str(ws.cell(row=r, column=2).value or "").split())
        acao = " ".join(str(ws.cell(row=r, column=4).value or "").split())
        st_  = " ".join(str(ws.cell(row=r, column=9).value or "").split())
        obs  = " ".join(str(ws.cell(row=r, column=10).value or "").split())
        rows.append({"n": n, "pilar": p, "acao": acao, "resp": resp,
                     "status": st_, "obs": obs})
    return pd.DataFrame(rows)

# ── Interface principal ───────────────────────────────────────────────────────
st.markdown("""
<div class="header-bar">
  <div>
    <div class="header-title">Plano de Monitoramento Previne Niterói — Biênio 2025–2026</div>
    <div class="header-sub">Secretaria Municipal de Fazenda · Pilares I, II, II.II e III</div>
  </div>
  <div style="text-align:right">
    <div class="header-badge">64 ações pactuadas</div>
    <div class="header-meta" style="margin-top:6px">Unidade de Controle Interno Setorial — UCIS/SMF</div>
  </div>
</div>
""", unsafe_allow_html=True)

# Upload
with st.container():
    uploaded = st.file_uploader(
        "📂 Carregue a planilha de monitoramento (.xlsx) para atualizar o dashboard",
        type=["xlsx"],
        help="Use o arquivo MODELO_PARA_MONITORAMENTO_2025-2026.xlsx"
    )

if uploaded is None:
    st.info("👆 Faça o upload da planilha acima para visualizar o dashboard atualizado.")
    st.stop()

# Ler dados
try:
    df = ler_planilha(uploaded)
except Exception as e:
    st.error(f"Erro ao ler a planilha: {e}\nVerifique se o arquivo é o modelo correto.")
    st.stop()

if df.empty:
    st.error("Nenhuma ação encontrada. Verifique se a planilha é o modelo correto.")
    st.stop()

# ── Métricas ──────────────────────────────────────────────────────────────────
total     = len(df)
concl     = (df.status == "Concluído").sum()
cont      = (df.status == "Contínuo").sum()
ea        = (df.status == "Em andamento").sum()
ni        = (df.status == "Não iniciado").sum()
entregue  = concl + cont
pct_e     = round(100 * entregue / total, 1)

# Cabeçalho owner
st.markdown("""
<div class="owner">
  <div class="lbl">Área detentora do dashboard</div>
  <div class="val">Unidade de Controle Interno Setorial — UCIS/SMF</div>
  <div class="desc">Responsável pelo acompanhamento das ações pactuadas do Previne Niterói,
  consolidação de evidências e reporte à Controladoria Geral do Município (CGM).</div>
</div>
""", unsafe_allow_html=True)

# Alerta
st.markdown(f"""
<div class="alerta">
  <b>Ponto de atenção:</b> o percentual entregue está em <b>{pct_e}%</b>
  ({entregue} de {total} ações). Restam <b>{ea + ni} pendências</b> para dez/2026 —
  {ea} em andamento{f" e {ni} não iniciada(s)" if ni > 0 else ", todas já em execução"}.
</div>
""", unsafe_allow_html=True)

# KPI cards
st.markdown(f"""
<div class="kpi-row">
  <div class="kpi" style="border-color:#0d3b6e">
    <div class="val">{total}</div><div class="lbl">ações pactuadas</div>
  </div>
  <div class="kpi" style="border-color:#2e7d32">
    <div class="val" style="color:#2e7d32">{pct_e}%</div>
    <div class="lbl">entregues ({entregue} de {total})</div>
  </div>
  <div class="kpi" style="border-color:#2e7d32">
    <div class="val" style="color:#2e7d32">{concl}</div>
    <div class="lbl">concluídas</div>
  </div>
  <div class="kpi" style="border-color:#1565c0">
    <div class="val" style="color:#1565c0">{cont}</div>
    <div class="lbl">contínuas</div>
  </div>
  <div class="kpi" style="border-color:#e8a100">
    <div class="val" style="color:#b57f00">{ea}</div>
    <div class="lbl">em andamento</div>
  </div>
  {"" if ni == 0 else f'<div class="kpi" style="border-color:#c62828"><div class="val" style="color:#c62828">{ni}</div><div class="lbl">não iniciadas</div></div>'}
</div>
""", unsafe_allow_html=True)

# ── Abas ─────────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs([
    "📊 Visão geral",
    f"⚠️ Em andamento ({ea})",
    "📋 Todas as ações",
])

# ══════════ ABA 1 — VISÃO GERAL ══════════════════════════════════════════════
with tab1:
    col1, col2 = st.columns(2)

    with col1:
        # Donut geral
        cnt = df.groupby("status").size().reset_index(name="qtd")
        fig_donut = px.pie(
            cnt, values="qtd", names="status", hole=.6,
            color="status", color_discrete_map=STATUS_COLORS,
            title="Distribuição geral de status"
        )
        fig_donut.update_traces(textinfo="value+percent", pull=[0.03]*len(cnt))
        fig_donut.update_layout(height=340, margin=dict(t=40,b=0,l=0,r=0),
                                legend=dict(orientation="h", y=-0.1))
        st.plotly_chart(fig_donut, use_container_width=True)

    with col2:
        # Barras empilhadas por pilar
        order = ["I", "II", "II.II", "III"]
        nomes = {
            "I": "Pilar I — Gestão Fazendária",
            "II": "Pilar II — Integridade",
            "II.II": "Pilar II.II — Alienação",
            "III": "Pilar III — Estruturantes",
        }
        status_list = ["Concluído", "Contínuo", "Em andamento", "Não iniciado"]
        fig_bar = go.Figure()
        for s in status_list:
            vals = [
                (df[(df.pilar == p) & (df.status == s)].shape[0]) for p in order
            ]
            fig_bar.add_trace(go.Bar(
                name=s, x=[nomes.get(p, p) for p in order], y=vals,
                marker_color=STATUS_COLORS.get(s, "#ccc"),
                text=vals, textposition="inside",
            ))
        fig_bar.update_layout(
            barmode="stack", height=340, title="Entregues por pilar",
            margin=dict(t=40,b=0,l=0,r=0),
            legend=dict(orientation="h", y=-0.15),
            xaxis_tickangle=-15,
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    # Série histórica
    meses_hist = list(SERIE_HIST.keys())
    vals_hist  = list(SERIE_HIST.values())
    meses_hist.append("ago/26")
    vals_hist.append(pct_e)

    fig_serie = go.Figure()
    fig_serie.add_trace(go.Scatter(
        x=meses_hist, y=vals_hist, mode="lines+markers",
        name="% atingido", line=dict(color="#1565c0", width=2.5),
        fill="tozeroy", fillcolor="rgba(21,101,192,.10)",
        marker=dict(size=7),
    ))
    fig_serie.add_trace(go.Scatter(
        x=meses_hist, y=[100]*len(meses_hist),
        mode="lines", name="Meta 100%",
        line=dict(color="#c62828", dash="dash", width=1.5),
    ))
    fig_serie.update_layout(
        title="Série histórica — percentual atingido × meta",
        height=300, margin=dict(t=40,b=0,l=0,r=0),
        yaxis=dict(range=[0,110], ticksuffix="%"),
        legend=dict(orientation="h", y=-0.2),
    )
    st.plotly_chart(fig_serie, use_container_width=True)

    # Cenários 80%
    st.subheader("Referência gerencial de 80%")
    meta80 = round(0.8 * total)
    faltam = max(0, meta80 - entregue)
    st.markdown(f"**80% de {total} ações = {meta80} entregues · posição atual: {entregue} · faltam {faltam} ações**")

    # Definir grupos de cenário com base nas obs
    df_pend = df[df.status == "Em andamento"].copy()
    pub  = df_pend[df_pend.obs.str.contains("aguardando publicação|minuta pronta|pendente de publicação", case=False, na=False)]
    val  = df_pend[df_pend.obs.str.contains("validar conclusão|validar|evidência", case=False, na=False)]
    cap  = df_pend[df_pend.obs.str.contains("capacitação|curso|ENAP|EGG|agendada", case=False, na=False)]

    cA = entregue + len(pub)
    cB = cA + len(val)
    cC = cB + len(cap)

    labels = [
        "Posição atual",
        f"+ {len(pub)} publicações",
        f"+ {len(val)} validações",
        f"+ {len(cap)} capacitações",
    ]
    valores = [
        round(100*entregue/total,1),
        round(100*cA/total,1),
        round(100*cB/total,1),
        round(100*cC/total,1),
    ]
    cores = ["#94a3b8", "#e8611a", "#e8a100", "#2e7d32"]

    fig_cen = go.Figure()
    fig_cen.add_trace(go.Bar(
        x=labels, y=valores, marker_color=cores,
        text=[f"{v}%" for v in valores], textposition="outside",
    ))
    fig_cen.add_hline(y=80, line_dash="dash", line_color="#c62828",
                      annotation_text="Meta 80%", annotation_position="right")
    fig_cen.update_layout(
        height=320, margin=dict(t=20,b=0,l=0,r=60),
        yaxis=dict(range=[0,105], ticksuffix="%"),
        showlegend=False,
    )
    st.plotly_chart(fig_cen, use_container_width=True)

    if len(pub) > 0:
        with st.expander(f"📄 Ver as {len(pub)} ações de publicação de minuta"):
            st.dataframe(pub[["n","pilar","acao","resp"]].rename(
                columns={"n":"Nº","pilar":"Pilar","acao":"Ação","resp":"Responsável"}),
                hide_index=True, use_container_width=True)
    if len(val) > 0:
        with st.expander(f"✅ Ver as {len(val)} ações de validação de evidência"):
            st.dataframe(val[["n","pilar","acao","resp"]].rename(
                columns={"n":"Nº","pilar":"Pilar","acao":"Ação","resp":"Responsável"}),
                hide_index=True, use_container_width=True)
    if len(cap) > 0:
        with st.expander(f"🎓 Ver as {len(cap)} ações de capacitação"):
            st.dataframe(cap[["n","pilar","acao","resp"]].rename(
                columns={"n":"Nº","pilar":"Pilar","acao":"Ação","resp":"Responsável"}),
                hide_index=True, use_container_width=True)

# ══════════ ABA 2 — EM ANDAMENTO ══════════════════════════════════════════════
with tab2:
    df_ea = df[df.status == "Em andamento"].copy()

    # Gráfico de categorias de gargalo (baseado nas obs)
    def categorizar(obs):
        obs = obs.lower()
        if any(k in obs for k in ["aguardando publicação","minuta","pendente de publicação"]):
            return "Processo parado e/ou aguardando movimentação"
        if any(k in obs for k in ["validar","evidência","conclusão"]):
            return "Validar conclusão / evidência"
        if any(k in obs for k in ["orçamento","orcamento"]):
            return "Restrição orçamentária"
        if any(k in obs for k in ["capacitação","curso","enap","egg","agendada"]):
            return "Capacitação pendente / agendada"
        if "normativo" in obs or "regulamento" in obs or "regimento" in obs:
            return "Tramitação normativa"
        if "pl " in obs or "projeto de lei" in obs or "legislativo" in obs:
            return "Decisão do Executivo (PL)"
        if obs.strip():
            return "Outro"
        return "Não informado"

    df_ea["categoria"] = df_ea["obs"].apply(categorizar)

    cat_colors = {
        "Processo parado e/ou aguardando movimentação": "#e8611a",
        "Validar conclusão / evidência": "#2e7d32",
        "Restrição orçamentária": "#c62828",
        "Capacitação pendente / agendada": "#1565c0",
        "Tramitação normativa": "#7c3aed",
        "Decisão do Executivo (PL)": "#0f766e",
        "Outro": "#b45309",
        "Não informado": "#cbd5e1",
    }

    col1, col2 = st.columns(2)
    with col1:
        cats = df_ea.groupby("categoria").size().reset_index(name="qtd").sort_values("qtd", ascending=True)
        fig_garg = px.bar(
            cats, x="qtd", y="categoria", orientation="h",
            color="categoria", color_discrete_map=cat_colors,
            title="Onde estão as travas",
            text="qtd",
        )
        fig_garg.update_layout(height=380, showlegend=False,
                               margin=dict(t=40,b=0,l=0,r=0))
        st.plotly_chart(fig_garg, use_container_width=True)

    with col2:
        por_pilar = df_ea.groupby("pilar").size().reset_index(name="qtd")
        fig_pilar = px.bar(
            por_pilar, x="pilar", y="qtd",
            color_discrete_sequence=["#e8a100"],
            title="Em andamento por pilar",
            text="qtd",
        )
        fig_pilar.update_layout(height=380, showlegend=False,
                                margin=dict(t=40,b=0,l=0,r=0))
        st.plotly_chart(fig_pilar, use_container_width=True)

    # Tabela das ações em andamento
    st.subheader(f"{len(df_ea)} ações em andamento")

    # Filtro por pilar
    pilares = ["Todos"] + sorted(df_ea.pilar.unique().tolist())
    pilar_sel = st.selectbox("Filtrar por pilar", pilares)
    df_show = df_ea if pilar_sel == "Todos" else df_ea[df_ea.pilar == pilar_sel]

    st.dataframe(
        df_show[["n","pilar","acao","resp","categoria"]].rename(columns={
            "n":"Nº","pilar":"Pilar","acao":"Ação pactuada",
            "resp":"Responsável","categoria":"Categoria do gargalo"
        }),
        hide_index=True, use_container_width=True, height=420,
    )

    # Exportar CSV
    csv = df_show[["n","pilar","acao","resp","obs","categoria"]].rename(columns={
        "n":"Nº","pilar":"Pilar","acao":"Ação pactuada",
        "resp":"Responsável","obs":"Observação","categoria":"Categoria do gargalo"
    }).to_csv(index=False, sep=";", encoding="utf-8-sig")
    st.download_button("⬇ Exportar CSV", csv, "em_andamento.csv", "text/csv")

# ══════════ ABA 3 — TODAS AS AÇÕES ═══════════════════════════════════════════
with tab3:
    # Filtros
    col1, col2, col3 = st.columns(3)
    with col1:
        f_pilar = st.selectbox("Pilar", ["Todos"] + sorted(df.pilar.unique().tolist()), key="f_pilar")
    with col2:
        f_status = st.selectbox("Status", ["Todos"] + list(STATUS_COLORS.keys()), key="f_status")
    with col3:
        f_busca = st.text_input("Buscar ação ou responsável", key="f_busca")

    df_f = df.copy()
    if f_pilar != "Todos":
        df_f = df_f[df_f.pilar == f_pilar]
    if f_status != "Todos":
        df_f = df_f[df_f.status == f_status]
    if f_busca:
        mask = (df_f.acao.str.contains(f_busca, case=False, na=False) |
                df_f.resp.str.contains(f_busca, case=False, na=False))
        df_f = df_f[mask]

    st.caption(f"{len(df_f)} ação(ões) exibida(s)")
    st.dataframe(
        df_f[["n","pilar","acao","resp","status"]].rename(columns={
            "n":"Nº","pilar":"Pilar","acao":"Ação pactuada",
            "resp":"Responsável","status":"Status"
        }),
        hide_index=True, use_container_width=True, height=500,
    )
